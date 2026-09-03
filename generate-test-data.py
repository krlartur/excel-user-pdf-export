"""
Generuje examples/dane_testowe.xlsx: 20 kolumn x 100 wierszy danych testowych
(arkusz "Dane") + gotowy przyklad konfiguracji (arkusz "Export_Config"),
uzywany do testowania macro EksportujRaportyUzytkownikow.
"""
import random
from datetime import date, timedelta

import openpyxl
from openpyxl.styles import Font

random.seed(42)

USERS = [
    "Jan Kowalski",
    "Anna Nowak",
    "Piotr Wisniewski",
    "Maria Wojcik",
    "Tomasz Kaminski",
    "Katarzyna Lewandowska",
    "Michal Zielinski",
    "Agnieszka Szymanska",
]

ZADANIA = [
    "Instalacja czujnika", "Przeglad okresowy", "Naprawa usterki",
    "Wymiana czesci", "Kalibracja urzadzenia", "Audyt bezpieczenstwa",
    "Szkolenie", "Konserwacja",
]
STATUSY = ["Zrobione", "W trakcie", "Oczekuje", "Anulowane"]
PRIORYTETY = ["Niski", "Sredni", "Wysoki", "Krytyczny"]
DZIALY = ["Produkcja", "Utrzymanie ruchu", "IT", "Logistyka", "Jakosc"]
LOKALIZACJE = ["Hala A", "Hala B", "Magazyn", "Biuro", "Zewnetrze"]
KLIENCI = ["ACME Sp. z o.o.", "Nowak-Tech", "BudMax", "Elektro-Plus", "LogiTrans"]
WYNIKI = ["OK", "Do poprawy", "Wymaga eskalacji"]

COLUMNS = [
    "Uzytkownik", "Data", "Zadanie", "Status", "Priorytet", "Dzial",
    "Lokalizacja", "Godziny", "Koszt", "Kategoria", "Klient",
    "Numer zlecenia", "Opis", "Uwagi", "Termin", "Wykonawca",
    "Zatwierdzone przez", "Data zatwierdzenia", "Wynik", "Notatki wewnetrzne",
]
assert len(COLUMNS) == 20

# Kolumny wybrane do przykladowego filtra w Export_Config (5 z 20)
KOLUMNY_DO_EKSPORTU = ["Data", "Zadanie", "Status", "Priorytet", "Klient"]

wb = openpyxl.Workbook()

# --- Arkusz "Dane" ---
ws = wb.active
ws.title = "Dane"
ws.append(COLUMNS)
for cell in ws[1]:
    cell.font = Font(bold=True)

start_date = date(2026, 1, 1)
for i in range(100):
    row_date = start_date + timedelta(days=random.randint(0, 240))
    approval_date = row_date + timedelta(days=random.randint(1, 10))
    row = [
        random.choice(USERS),                                  # Uzytkownik
        row_date,                                               # Data
        random.choice(ZADANIA),                                 # Zadanie
        random.choice(STATUSY),                                 # Status
        random.choice(PRIORYTETY),                               # Priorytet
        random.choice(DZIALY),                                   # Dzial
        random.choice(LOKALIZACJE),                              # Lokalizacja
        round(random.uniform(0.5, 8), 1),                        # Godziny
        round(random.uniform(50, 5000), 2),                      # Koszt
        random.choice(["Serwis", "Inwestycja", "Reklamacja"]),   # Kategoria
        random.choice(KLIENCI),                                  # Klient
        f"ZL-{2026}-{1000 + i}",                                 # Numer zlecenia
        f"Opis zgloszenia nr {i + 1}",                           # Opis
        "" if i % 3 else "Wymaga dodatkowej weryfikacji",        # Uwagi
        row_date + timedelta(days=random.randint(1, 14)),        # Termin
        random.choice(USERS),                                    # Wykonawca
        random.choice(USERS),                                    # Zatwierdzone przez
        approval_date,                                            # Data zatwierdzenia
        random.choice(WYNIKI),                                    # Wynik
        f"Notatka wewnetrzna #{i + 1}",                           # Notatki wewnetrzne
    ]
    ws.append(row)

for col_cells in ws.columns:
    max_len = max(len(str(c.value)) if c.value is not None else 0 for c in col_cells)
    ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 2, 28)

# --- Arkusz "Export_Config" ---
cfg = wb.create_sheet("Export_Config")
cfg["A1"] = "Nazwa arkusza z danymi:"
cfg["B1"] = "Dane"
cfg["A2"] = "Kolumna z uzytkownikami:"
cfg["B2"] = "Uzytkownik"
cfg["A3"] = "Folder docelowy na PDF:"
cfg["B3"] = ""  # user wypelnia sam, albo makro zapyta o folder
cfg["A4"] = "Prefiks nazwy pliku (opcjonalny):"
cfg["B4"] = "Raport"

cfg["A6"] = "Uzytkownicy do eksportu"
cfg["C6"] = "Kolumny do eksportu"
for cell in (cfg["A6"], cfg["C6"]):
    cell.font = Font(bold=True)

przykladowi_uzytkownicy = ["Jan Kowalski", "Anna Nowak", "Piotr Wisniewski"]
for idx, u in enumerate(przykladowi_uzytkownicy):
    cfg.cell(row=7 + idx, column=1, value=u)

for idx, h in enumerate(KOLUMNY_DO_EKSPORTU):
    cfg.cell(row=7 + idx, column=3, value=h)

for col, width in (("A", 22), ("B", 40), ("C", 22)):
    cfg.column_dimensions[col].width = width

wb.save("dane_testowe.xlsx")
print("Zapisano dane_testowe.xlsx")
